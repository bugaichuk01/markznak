"""Генерация шаблонов Excel и импорт связок GTIN ↔ OZON ID / кодов маркировки."""

from __future__ import annotations

import csv
import io
import re
import uuid
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models import EmissionOrder, EmissionOrderStatus, OzonMapping
from schemas import ExcelTemplateProduct


EXCEL_COLS = ("Артикул", "Наименование", "GTIN", "OZON ID")


def _norm_gtin(raw: str) -> str:
    digits = re.sub(r"\D", "", (raw or "").strip())
    if not digits:
        return ""
    if len(digits) < 8:
        return ""
    return digits[:14]


def _normalize_header(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def generate_template_bytes(items: list[ExcelTemplateProduct]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон"
    for col, title in enumerate(EXCEL_COLS, start=1):
        ws.cell(row=1, column=col, value=title)
        ws.column_dimensions[get_column_letter(col)].width = 22
    for i, item in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=item.article)
        ws.cell(row=i, column=2, value=item.name)
        ws.cell(row=i, column=3, value=item.gtin)
        ws.cell(row=i, column=4, value=item.ozon_id or "")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_map(headers: list[str]) -> dict[str, int]:
    """Имя колонки (нормализованное) → индекс."""
    m: dict[str, int] = {}
    for idx, h in enumerate(headers):
        key = _normalize_header(h)
        if key:
            m[key] = idx
    return m


def _col_index(header_map: dict[str, int], *aliases: str) -> int | None:
    for a in aliases:
        k = _normalize_header(a)
        if k in header_map:
            return header_map[k]
    return None


def _read_rows_xlsx(data: bytes) -> tuple[list[str], list[list[str | None]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    body: list[list[str | None]] = []
    for row in rows_iter:
        body.append([str(c).strip() if c is not None else None for c in row])
    return headers, body


def _read_rows_csv(data: bytes) -> tuple[list[str], list[list[str | None]]]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        return [], []
    headers = [c.strip() for c in rows[0]]
    body: list[list[str | None]] = []
    for r in rows[1:]:
        body.append([c.strip() if c else None for c in r])
    return headers, body


async def import_ozon_ids_from_file(
    session: AsyncSession,
    filename: str,
    file_body: bytes,
) -> tuple[int, int, int]:
    """
    Читает GTIN и OZON ID из первого листа Excel или CSV.
    Пустые строки пропускаются. Новые GTIN получают placeholder name/article.
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        headers, body = _read_rows_csv(file_body)
    elif lower.endswith(".xlsx"):
        headers, body = _read_rows_xlsx(file_body)
    else:
        raise ValueError("Ожидается файл .xlsx или .csv")

    if not headers:
        raise ValueError("Пустой или некорректный файл: нет строки заголовков")

    hm = _header_map(headers)
    idx_gtin = _col_index(hm, "GTIN", "gtin", "гтин", "штрихкод", "EAN")
    idx_ozon = _col_index(hm, "OZON ID", "ozon id", "ozon_id", "озон id", "id ozon")
    if idx_gtin is None or idx_ozon is None:
        raise ValueError('В файле должны быть колонки "GTIN" и "OZON ID"')

    pairs: dict[str, str] = {}
    skipped = 0

    for row in body:
        if idx_gtin >= len(row) or idx_ozon >= len(row):
            skipped += 1
            continue
        gtin_raw = row[idx_gtin] or ""
        ozon_raw = row[idx_ozon] or ""
        if not ozon_raw.strip():
            skipped += 1
            continue
        gtin = _norm_gtin(gtin_raw)
        if not gtin:
            skipped += 1
            continue
        ozon_id = ozon_raw.strip()[:64]
        pairs[gtin] = ozon_id

    created = 0
    updated = 0

    for gtin, ozon_id in pairs.items():
        existing = await session.scalar(select(OzonMapping).where(OzonMapping.gtin == gtin))
        if existing:
            existing.ozon_id = ozon_id
            updated += 1
        else:
            session.add(
                OzonMapping(
                    gtin=gtin,
                    article=gtin,
                    name="Импорт OZON ID",
                    ozon_id=ozon_id,
                )
            )
            created += 1

    await session.commit()
    return created, updated, skipped


def _read_marking_codes_csv(file_body: bytes) -> list[str]:
    text = file_body.decode("utf-8-sig", errors="replace")
    codes: list[str] = []
    # GS1-разделители (\x1d, \x1e) в CSV часто превращаются в переносы строк — собираем части в один код
    current_code = ""
    for line in text.splitlines():
        line = line.strip().strip('"').strip("'")
        if not line:
            continue
        if line.startswith("01") and len(line) > 10:
            if current_code:
                codes.append(current_code)
            current_code = line
        elif current_code:
            current_code += line
        else:
            codes.append(line)
    if current_code:
        codes.append(current_code)
    return codes


def _read_marking_codes_xlsx(file_body: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(file_body), read_only=True, data_only=True)
    ws = wb.active
    codes: list[str] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and row[0] is not None:
            code = str(row[0]).strip()
            if code:
                codes.append(code)
    wb.close()
    return codes


async def import_marking_codes_from_file(
    session: AsyncSession,
    filename: str,
    file_body: bytes,
) -> tuple[int, int, list[str]]:
    """
    Импорт кодов маркировки из CSV или Excel.
    Возвращает (добавлено, пропущено, ошибки).

    CSV: один код на строку.
    Excel (.xlsx): коды в первом столбце.
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        codes = _read_marking_codes_csv(file_body)
    elif lower.endswith(".xlsx"):
        codes = _read_marking_codes_xlsx(file_body)
    else:
        raise ValueError("Поддерживаются только .csv и .xlsx файлы")

    if not codes:
        raise ValueError("Файл не содержит кодов маркировки")

    import_order = EmissionOrder(
        id=uuid.uuid4(),
        gtin=None,
        quantity=len(codes),
        status=EmissionOrderStatus.AVAILABLE,
        suz_order_id=None,
        suz_marking_codes=codes,
    )
    session.add(import_order)
    await session.commit()

    return len(codes), 0, []
