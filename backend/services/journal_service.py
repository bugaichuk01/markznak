"""Сервис журнала операций."""
from __future__ import annotations

import io
import logging
from datetime import datetime
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from models import OperationLog, OperationLogStatus, OperationLogType

logger = logging.getLogger(__name__)


async def log_operation(
    db: AsyncSession,
    operation_type: OperationLogType,
    status: OperationLogStatus = OperationLogStatus.SUCCESS,
    description: str | None = None,
    related_id: str | None = None,
    related_type: str | None = None,
    codes_count: int | None = None,
    gtin: str | None = None,
    error_message: str | None = None,
    details: dict | None = None,
    org_id=None,
) -> OperationLog | None:
    """Записать операцию в журнал."""
    try:
        entry = OperationLog(
            operation_type=operation_type,
            status=status,
            description=description,
            related_id=str(related_id) if related_id else None,
            related_type=related_type,
            codes_count=codes_count,
            gtin=gtin,
            error_message=error_message,
            details=details,
            org_id=org_id,
        )
        db.add(entry)
        await db.commit()
        await db.refresh(entry)
        return entry
    except Exception as e:
        logger.error("Ошибка записи в журнал: %s", e)
        return None


async def get_journal(
    db: AsyncSession,
    operation_type: str | None = None,
    status: str | None = None,
    gtin: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    search: str | None = None,
    limit: int = 50,
    offset: int = 0,
    org_id=None,
) -> tuple[list[OperationLog], int]:
    """Получить список операций с фильтрацией."""
    filters = []
    if org_id:
        filters.append(OperationLog.org_id == org_id)
    if operation_type:
        filters.append(OperationLog.operation_type == operation_type)
    if status:
        filters.append(OperationLog.status == status)
    if gtin:
        filters.append(OperationLog.gtin == gtin)
    if date_from:
        filters.append(OperationLog.created_at >= date_from)
    if date_to:
        filters.append(OperationLog.created_at <= date_to)
    if search:
        filters.append(
            or_(
                OperationLog.description.ilike(f"%{search}%"),
                OperationLog.gtin.ilike(f"%{search}%"),
                OperationLog.related_id.ilike(f"%{search}%"),
            )
        )

    query = select(OperationLog)
    count_query = select(func.count()).select_from(OperationLog)
    if filters:
        query = query.where(and_(*filters))
        count_query = count_query.where(and_(*filters))

    total = await db.scalar(count_query) or 0
    result = await db.scalars(
        query.order_by(OperationLog.created_at.desc()).limit(limit).offset(offset)
    )
    return list(result.all()), total


async def export_journal_excel(
    db: AsyncSession,
    **filters,
) -> bytes:
    """Экспорт журнала в Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    entries, _ = await get_journal(db, limit=10000, offset=0, **filters)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал операций"

    headers = [
        "Дата и время",
        "Тип операции",
        "Статус",
        "Описание",
        "GTIN",
        "Кол-во кодов",
        "ID документа",
        "Ошибка",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"
        )

    type_labels = {
        "order_created": "Создан заказ СУЗ",
        "order_sent": "Заказ отправлен",
        "codes_downloaded": "КМ скачаны",
        "order_closed": "Заказ закрыт",
        "utilisation_sent": "Ввод в оборот",
        "withdrawal_sent": "Вывод из оборота",
        "aggregation_sent": "Агрегация КИТУ",
        "return_sent": "Возврат в оборот",
        "upd_created": "УПД создан",
        "upd_sent": "УПД отправлен",
        "cis_checked": "Проверка статуса КМ",
        "label_printed": "Печать этикеток",
        "card_created": "Карточка НК",
        "token_updated": "Токен обновлён",
    }
    status_labels = {
        "success": "Успешно",
        "error": "Ошибка",
        "pending": "В обработке",
    }

    for row_num, entry in enumerate(entries, 2):
        op_type = (
            entry.operation_type.value
            if hasattr(entry.operation_type, "value")
            else str(entry.operation_type)
        )
        st = (
            entry.status.value
            if hasattr(entry.status, "value")
            else str(entry.status)
        )
        ws.cell(
            row=row_num,
            column=1,
            value=entry.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        )
        ws.cell(row=row_num, column=2, value=type_labels.get(op_type, op_type))
        ws.cell(row=row_num, column=3, value=status_labels.get(st, st))
        ws.cell(row=row_num, column=4, value=entry.description or "")
        ws.cell(row=row_num, column=5, value=entry.gtin or "")
        ws.cell(row=row_num, column=6, value=entry.codes_count or "")
        ws.cell(row=row_num, column=7, value=entry.related_id or "")
        ws.cell(row=row_num, column=8, value=entry.error_message or "")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 38
    ws.column_dimensions["H"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
