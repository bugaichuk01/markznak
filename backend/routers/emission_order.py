import csv
import io
import uuid as uuid_lib
from uuid import UUID

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl.styles import Font, PatternFill
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from database import get_db_session
from dependencies import get_current_org, get_current_user
from models import EmissionOrder, EmissionOrderStatus, Organization, User
from schemas import (
    CisStatusItem,
    CisStatusRequest,
    CisStatusResponse,
    CloseOrderRequest,
    CloseOrderResponse,
    EmissionOrderCreate,
    EmissionOrderGtinPatch,
    EmissionOrderResponse,
    EmissionOrderStatusUpdateRequest,
    FetchCodesResponse,
    IntroduceOstBodyPreview,
    IntroduceOstBodyRequest,
    IntroduceOstRequest,
    MarkingCodePrintOptionsResponse,
    MarkingCodesListResponse,
    MergeOrdersRequest,
    SuzConnectivityDiagnosticsResponse,
    SuzCreateOrderProxyRequest,
    SuzOrderPayloadPreview,
    SuzSendOrderPayload,
    SuzSendOrderRequest,
    SuzSendOrderResponse,
    SuzSyncResponse,
)
from services import emission_order_service
from services.suz_diagnostic_service import diagnose_suz_oms_endpoint
from models import OperationLogStatus, OperationLogType
from services.journal_service import log_operation
from services.suz_integration_service import check_cis_statuses_batch
router = APIRouter(prefix="/emission-orders", tags=["SUZ"])
@router.post("/", response_model=EmissionOrderResponse, status_code=status.HTTP_201_CREATED)
async def create_emission_order(
    data: EmissionOrderCreate,
    _: User = Depends(get_current_user),
    org: Organization | None = Depends(get_current_org),
    db: AsyncSession = Depends(get_db_session),
) -> EmissionOrderResponse:
    return await emission_order_service.create_order(
        data, db, org_id=org.id if org else None
    )
@router.get("/", response_model=list[EmissionOrderResponse])
async def list_emission_orders(
    _: User = Depends(get_current_user),
    org: Organization | None = Depends(get_current_org),
    db: AsyncSession = Depends(get_db_session),
) -> list[EmissionOrderResponse]:
    return await emission_order_service.get_orders(db, org_id=org.id if org else None)
@router.get("/marking-codes-for-print", response_model=MarkingCodePrintOptionsResponse)
async def list_marking_codes_for_print(
    _: User = Depends(get_current_user),
    org: Organization | None = Depends(get_current_org),
    db: AsyncSession = Depends(get_db_session),
) -> MarkingCodePrintOptionsResponse:
    codes = await emission_order_service.list_marking_codes_for_print(
        db, org_id=org.id if org else None
    )
    return MarkingCodePrintOptionsResponse(codes=codes)
@router.get("/codes", response_model=MarkingCodesListResponse)
async def list_marking_codes(
    gtin: str | None = None,
    order_id: str | None = None,
    search: str | None = None,
    limit: int = 100,
    offset: int = 0,
    _: User = Depends(get_current_user),
    org: Organization | None = Depends(get_current_org),
    db: AsyncSession = Depends(get_db_session),
) -> MarkingCodesListResponse:
    data = await emission_order_service.list_all_marking_codes(
        db,
        gtin=gtin,
        order_id=order_id,
        search=search,
        limit=limit,
        offset=offset,
        org_id=org.id if org else None,
    )
    return MarkingCodesListResponse(**data)
@router.post("/codes/check-status", response_model=CisStatusResponse)
async def check_codes_status(
    request: CisStatusRequest,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> CisStatusResponse:
    if len(request.cises) > 50:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Максимум 50 кодов за один запрос",
        )
    results = await check_cis_statuses_batch(request.cises, db=db)
    items = [
        CisStatusItem(
            cis=r["cis"],
            status=r.get("status"),
            owner_inn=r.get("owner_inn"),
            owner_name=r.get("owner_name"),
            gtin=r.get("gtin"),
            produced_date=r.get("produced_date"),
            error=r.get("error"),
        )
        for r in results
    ]
    await log_operation(
        db,
        operation_type=OperationLogType.CIS_STATUS_CHECKED,
        status=OperationLogStatus.SUCCESS,
        description=f"Проверен статус {len(request.cises)} кодов",
        codes_count=len(request.cises),
    )
    return CisStatusResponse(
        results=items,
        total=len(request.cises),
        checked=len([r for r in results if "error" not in r]),
    )


@router.post("/codes/import-csv")
async def import_codes_from_csv(
    file: UploadFile = File(...),
    _: User = Depends(get_current_user),
    org: Organization | None = Depends(get_current_org),
    db: AsyncSession = Depends(get_db_session),
) -> dict:
    """Импорт кодов маркировки из CSV/TXT файла (один код на строку)."""
    if not file.filename:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Файл не выбран")

    content = await file.read()
    text = content.decode("utf-8", errors="replace")

    lines: list[str] = []
    for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = line.strip().strip(";").strip()
        if line and len(line) >= 20:
            line = line.replace("\\x1d", "\x1d").replace("\\u001d", "\x1d")
            lines.append(line)

    if not lines:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Файл не содержит кодов маркировки")
    if len(lines) > 10000:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Максимум 10000 кодов за один импорт")

    order = EmissionOrder(
        gtin=None,
        quantity=len(lines),
        status=EmissionOrderStatus.CLOSED,
        suz_order_id=f"IMPORT-{str(uuid_lib.uuid4())[:8]}",
        suz_marking_codes=lines,
        org_id=org.id if org else None,
    )
    db.add(order)
    await db.commit()
    await db.refresh(order)

    return {
        "success": True,
        "imported": len(lines),
        "order_id": str(order.id),
        "preview": [c[:40] + "..." for c in lines[:3]],
    }


@router.get("/codes/export-csv")
async def export_codes_to_csv(
    gtin: str | None = Query(None),
    order_id: str | None = Query(None),
    _: User = Depends(get_current_user),
    org: Organization | None = Depends(get_current_org),
    db: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    """Экспорт кодов маркировки в CSV с фильтром по GTIN или order_id."""
    query = select(EmissionOrder)
    if org:
        query = query.where(EmissionOrder.org_id == org.id)
    if gtin:
        query = query.where(EmissionOrder.gtin == gtin)
    if order_id:
        query = query.where(EmissionOrder.id == UUID(order_id))

    orders = (await db.scalars(query)).all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["code", "gtin", "order_id", "status"])

    total = 0
    for order in orders:
        status_value = order.status.value if hasattr(order.status, "value") else str(order.status)
        for code in order.suz_marking_codes or []:
            writer.writerow([code, order.gtin or "", str(order.id), status_value])
            total += 1

    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=codes_export_{total}.csv",
        },
    )


@router.post("/import-excel-orders")
async def import_orders_from_excel(
    file: UploadFile = File(...),
    _: User = Depends(get_current_user),
    org: Organization | None = Depends(get_current_org),
    db: AsyncSession = Depends(get_db_session),
) -> dict:
    """Импорт заказов КМ из Excel. Колонки: GTIN, Количество, Товарная группа, Способ выпуска."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Только .xlsx или .xls файлы",
        )

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл пустой или нет данных после заголовка",
        )

    results: list[dict] = []
    errors: list[str] = []

    for idx, row in enumerate(rows, start=2):
        try:
            if not row or not row[0]:
                continue

            gtin = str(row[0]).strip().zfill(14) if row[0] else None
            quantity = int(row[1]) if len(row) > 1 and row[1] else 1
            product_group = (
                str(row[2]).strip().lower() if len(row) > 2 and row[2] else "perfumery"
            )
            release_method = (
                str(row[3]).strip().upper() if len(row) > 3 and row[3] else "PRODUCTION"
            )

            if not gtin or len(gtin) != 14:
                errors.append(f"Строка {idx}: неверный GTIN '{row[0]}'")
                continue
            if quantity < 1 or quantity > 150000:
                errors.append(f"Строка {idx}: неверное количество '{row[1]}'")
                continue

            product_card = await emission_order_service.resolve_product_card_by_gtin(db, gtin)

            order = EmissionOrder(
                gtin=gtin,
                quantity=quantity,
                status=EmissionOrderStatus.CREATED,
                release_method_type=release_method,
                product_card_id=product_card.id if product_card else None,
                org_id=org.id if org else None,
            )
            db.add(order)
            await db.flush()

            results.append({
                "row": idx,
                "order_id": str(order.id),
                "gtin": gtin,
                "quantity": quantity,
                "product_group": product_group,
                "release_method": release_method,
                "status": "created",
            })

        except Exception as e:
            errors.append(f"Строка {idx}: {str(e)}")

    await db.commit()

    return {
        "created": len(results),
        "errors": errors,
        "orders": results,
    }


@router.get("/excel-template")
async def download_order_template() -> Response:
    """Скачать шаблон Excel для импорта заказов."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы КМ"

    headers = ["GTIN (14 цифр)", "Количество", "Товарная группа", "Способ выпуска"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    examples = [
        ["02900004064948", 10, "perfumery", "PRODUCTION"],
        ["04601234567890", 5, "clothes", "REMAINS"],
        ["04600000000000", 20, "shoes", "REMARK"],
    ]
    for row_idx, row in enumerate(examples, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws.cell(row=6, column=1, value="Товарные группы:")
    ws.cell(row=7, column=1, value="perfumery, clothes, shoes, linen, milk, water, tires")
    ws.cell(row=9, column=1, value="Способы выпуска:")
    ws.cell(row=10, column=1, value="PRODUCTION, REMAINS, REMARK, REAPPLY")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=order_template.xlsx"},
    )


@router.post("/introduce-ost-body", response_model=IntroduceOstBodyPreview)
async def get_introduce_ost_body(
    data: IntroduceOstBodyRequest,
    _: User = Depends(get_current_user),
) -> IntroduceOstBodyPreview:
    """Получить тело документа LP_INTRODUCE_OST для подписи."""
    from services.remains_service import encode_introduce_ost_body

    if not data.marking_codes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Список кодов пуст",
        )

    body_json, body_b64 = encode_introduce_ost_body(data.marking_codes, data.product_group)
    return IntroduceOstBodyPreview(body=body_json, body_b64=body_b64)


@router.post("/introduce-ost")
async def send_introduce_ost_endpoint(
    data: IntroduceOstRequest,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> dict:
    """Отправить LP_INTRODUCE_OST — ввод в оборот остатков."""
    from services.remains_service import send_introduce_ost as _send

    if not data.marking_codes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Список кодов пуст",
        )

    try:
        return await _send(
            data.marking_codes,
            data.signature,
            data.product_group,
            db,
        )
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(e)) from e
    except RuntimeError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e)) from e


@router.get("/diagnostics/connectivity", response_model=SuzConnectivityDiagnosticsResponse)
async def suz_connectivity_diagnostics(
    _: User = Depends(get_current_user),
) -> SuzConnectivityDiagnosticsResponse:
    data = await diagnose_suz_oms_endpoint()
    return SuzConnectivityDiagnosticsResponse.model_validate(data)
@router.post("/create", response_model=SuzSendOrderResponse)
async def create_suz_order_proxy(
    data: SuzCreateOrderProxyRequest,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> SuzSendOrderResponse:
    emission_order, suz_remote_id, raw_payload = await emission_order_service.create_suz_order_via_proxy(
        body_string=data.body_string,
        signature=data.signature,
        db=db,
        local_order_id=data.local_order_id,
    )
    return SuzSendOrderResponse(
        emission_order=(
            EmissionOrderResponse.model_validate(emission_order) if emission_order is not None else None
        ),
        suz=SuzSendOrderPayload(remote_order_id=suz_remote_id, payload=raw_payload),
    )
@router.post("/sync-from-suz", response_model=SuzSyncResponse)
async def sync_emission_orders_from_suz(
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> SuzSyncResponse:
    data = await emission_order_service.sync_orders_from_suz(db)
    return SuzSyncResponse(**data)
@router.post("/merge", response_model=EmissionOrderResponse, status_code=status.HTTP_201_CREATED)
async def merge_emission_orders(
    data: MergeOrdersRequest,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> EmissionOrderResponse:
    return await emission_order_service.merge_orders(data.order_ids, db)
@router.patch("/{order_id}/status", response_model=EmissionOrderResponse)
async def patch_emission_order_status(
    order_id: UUID,
    data: EmissionOrderStatusUpdateRequest,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> EmissionOrderResponse:
    return await emission_order_service.update_order_status(order_id, data.status.value, db)
@router.patch("/{order_id}/gtin", response_model=EmissionOrderResponse)
async def patch_emission_order_gtin(
    order_id: UUID,
    data: EmissionOrderGtinPatch,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> EmissionOrderResponse:
    return await emission_order_service.patch_order_gtin(order_id, data.gtin, db)
@router.get("/{order_id}/suz-order-payload", response_model=SuzOrderPayloadPreview)
async def get_suz_order_send_payload(
    order_id: UUID,
    release_method_type: str | None = None,
    producer: str | None = None,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> SuzOrderPayloadPreview:
    data = await emission_order_service.prepare_suz_order_send_payload(
        order_id,
        db,
        release_method_type=release_method_type,
        producer=producer,
    )
    return SuzOrderPayloadPreview.model_validate(data)
@router.post("/{order_id}/send", response_model=SuzSendOrderResponse)
async def send_emission_order_to_suz(
    order_id: UUID,
    data: SuzSendOrderRequest,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> SuzSendOrderResponse:
    emission_order, suz_remote_id, raw_payload = await emission_order_service.send_order_to_suz(
        order_id,
        db,
        body_string=data.body_string,
        signature=data.signature,
    )
    return SuzSendOrderResponse(
        emission_order=EmissionOrderResponse.model_validate(emission_order),
        suz=SuzSendOrderPayload(remote_order_id=suz_remote_id, payload=raw_payload),
    )
@router.post("/{order_id}/fetch-codes", response_model=FetchCodesResponse)
async def fetch_codes_from_suz(
    order_id: UUID,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> FetchCodesResponse:
    order, codes = await emission_order_service.download_order_codes(order_id, db)
    return FetchCodesResponse(
        order_id=order_id,
        codes_count=len(codes),
        status=order.status.value,
    )
@router.post("/{order_id}/close", response_model=CloseOrderResponse)
async def close_emission_order(
    order_id: UUID,
    data: CloseOrderRequest,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> CloseOrderResponse:
    order = await emission_order_service.close_order(
        order_id, db, signature=data.signature
    )
    return CloseOrderResponse(
        success=True,
        order_id=str(order_id),
        status=order.status.value,
    )
@router.get("/{order_id}/codes.csv")
async def download_codes_csv(
    order_id: UUID,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db_session),
) -> Response:
    order = await db.get(EmissionOrder, order_id)
    if order is None or not order.suz_marking_codes:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Коды не найдены")
    content = "\n".join(order.suz_marking_codes)
    return Response(
        content=content,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=codes_{order_id}.csv"
        },
    )
